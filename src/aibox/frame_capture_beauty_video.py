from pathlib import Path

import cv2
import os
import numpy as np

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

import torch
import torch.nn as nn
import clip
from PIL import Image

from aibox.test import EXTS

# 视频路径与输出目录
video_path = r"C:\Users\pablozhao\Desktop\Tubedown Download\1920x1080-30fps_POUSv49V.mp4"
output_dir = "frames_1080P_1.5x"
os.makedirs(output_dir, exist_ok=True)
excel_output_path = "high_quality_frames_1080P_1.5x.xlsx"


class AestheticPredictorV2(nn.Module):
    def __init__(self, input_dim=768):
        super().__init__()
        self.input_dim = input_dim
        self.layers = nn.Sequential(
            nn.Linear(self.input_dim, 1024),
            nn.Dropout(0.2),
            nn.Linear(1024, 128),
            nn.Dropout(0.2),
            nn.Linear(128, 64),
            nn.Dropout(0.1),
            nn.Linear(64, 16),
            nn.Linear(16, 1)
        )

    def forward(self, x):
        return self.layers(x)


WEIGHT_FILE = "ava+logos-l14-linearMSE.pth"
device = "cuda" if torch.cuda.is_available() else "cpu"
print("正在载入 OpenAI CLIP ViT-L/14 ...")
clip_model, clip_preprocess = clip.load("ViT-L/14", device=device)
model_aesthetic = AestheticPredictorV2(768).to(device)
model_aesthetic.load_state_dict(torch.load(WEIGHT_FILE, map_location=device))
model_aesthetic.eval()


def get_laion_v2_score(cv2_frame):
    """ 传入本地 OpenCV 读取的单帧图像，输出 LAION V2 美学打分 """
    rgb_frame = cv2.cvtColor(cv2_frame, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(rgb_frame)
    image_tensor = clip_preprocess(pil_img).unsqueeze(0).to(device)

    with torch.no_grad():
        image_features = clip_model.encode_image(image_tensor)
        image_features /= image_features.norm(dim=-1, keepdim=True)
        prediction = model_aesthetic(image_features.float())

    return prediction.item()


def get_video_info(video_path: str) -> dict:
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise ValueError(f"无法打开视频文件: {video_path}")
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS)
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    resolution_ratio = f"{width}*{height}"

    info = {
        "width": width,
        "height": height,
        "fps": fps,
        "frame_count": frame_count,
        "resolution_ratio": resolution_ratio,
    }
    if info["fps"] > 0:
        info["duration"] = round(info["frame_count"] / info["fps"], 2)
    cap.release()
    return info


# 判断是否为高质量帧（修改：返回是否通过及详细数据）
def check_frame_quality(frame, prev_frame=None):
    metrics = {
        "sharpness": 0.0,
        "brightness": 0.0,
        "mse_diff": None,
        "aesthetic_score": None,
        "remark": "通过验证"
    }

    # 1. 检查清晰度
    sharp = frame_sharpness(frame)
    metrics["sharpness"] = round(sharp, 2)
    if sharp < 120:
        metrics["remark"] = f"未通过: 清晰度过低 ({sharp:.2f} < 阈值: 120)"
        print(f"       [未通过] 原因: 清晰度过低 (当前值: {sharp:.2f} < 阈值: 120) -> 画面可能模糊/抖动")
        return False, metrics

    # 2. 检查亮度
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    brightness = gray.mean()
    metrics["brightness"] = round(brightness, 2)
    if brightness < 40 or brightness > 220:
        metrics["remark"] = f"未通过: 曝光异常 (亮度: {brightness:.2f})"
        print(f"       [未通过] 原因: 曝光异常 (当前亮度: {brightness:.2f}, 正常范围: 40~220)")
        return False, metrics

    # 3. 画面重复度过滤
    if prev_frame is not None:
        diff = frame_diff_roi(prev_frame, frame)
        metrics["mse_diff"] = round(diff, 2)
        if diff < 1500:
            metrics["remark"] = f"未通过: 与前帧太相似 (MSE: {diff:.2f} < 1500)"
            print(f"       [未通过] 与上一帧太相似 (ROI帧差 MSE: {diff:.2f} < 阈值: 1500)")
            return False, metrics
        else:
            print(f"       [检查通过] 画面有足够物理位移 (ROI帧差 MSE: {diff:.2f})")

    # 4. 运行美学评分
    aesthetic_score = get_laion_v2_score(frame)
    metrics["aesthetic_score"] = round(aesthetic_score, 3)
    print(f"       [美学得分]  : {aesthetic_score:.3f}")


    dr = estimate_dynamic_range(frame)
    if 175 <= dr <= 235:
        print(f"       [检查通过] 画面动态范围合适 (HDR: {dr:.2f})")
    else:
        print(f"       [未通过] 画面动态范围过低或过高 ，照片容易死黑或死白过平 (HDR: {dr:.2f})")
        return False, metrics
    noise = estimate_noise_std(frame)
    metrics["noise_std"] = round(noise, 2)
    if noise > 5.0:
        metrics["remark"] = f"未通过: 噪点过高 ({noise:.2f})"
        return False, metrics

    # 根据分值附加美学档次备注
    if aesthetic_score >= 6.0:
        metrics["remark"] = f"优秀帧 (美学分: {metrics['aesthetic_score']})"
    else:
        metrics["remark"] = "常规高质量帧"

    print(f"       [通过验证] (清晰度: {sharp:.2f} | 亮度: {brightness:.2f})")
    return True, metrics


def frame_sharpness(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    return cv2.Laplacian(gray, cv2.CV_64F).var()


def frame_diff_roi(img1, img2):
    h, w, _ = img1.shape
    start_y = int(h * 0.2)
    end_y = int(h * 0.7)
    roi1 = img1[start_y:end_y, :]
    roi2 = img2[start_y:end_y, :]
    g1 = cv2.cvtColor(roi1, cv2.COLOR_BGR2GRAY)
    g2 = cv2.cvtColor(roi2, cv2.COLOR_BGR2GRAY)
    return np.mean((g1.astype(float) - g2.astype(float)) ** 2)


def estimate_dynamic_range(
    img_bgr,
    percentile_low=1,
    percentile_high=99
) -> float:
    """
    基于灰度直方图的动态范围估计（0~255）
    数值越大，动态范围越高
    """
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)

    p_low = np.percentile(gray, percentile_low)
    p_high = np.percentile(gray, percentile_high)

    return float(p_high - p_low)


def estimate_noise_std(img_bgr):
    """
    基于平滑区域估计亮度噪声 (Y通道)
    返回近似噪声 σ
    """
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY).astype(np.float32)

    # Laplacian 用来找“非边缘 / 平滑区域”
    lap = cv2.Laplacian(gray, cv2.CV_32F)
    edge_strength = np.abs(lap)

    # 取平滑区域（Laplacian 较小）
    smooth_mask = edge_strength < np.percentile(edge_strength, 30)
    smooth_vals = gray[smooth_mask]

    if smooth_vals.size < 100:
        # fallback：整图 std
        return float(np.std(gray))

    # 局部窗口 噪声估计
    noise_map = cv2.GaussianBlur(gray, (3, 3), 0)
    noise_residual = gray - noise_map

    noise_std = float(np.std(noise_residual[smooth_mask]))
    return noise_std


# 导出并美化 Excel 的函数
def save_to_styled_excel(data_list, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "高质量关键帧"

    # 启用网格线显示
    ws.views.sheetView[0].showGridLines = True

    # 表头定义
    headers = ["帧序号", "保存路径", "时间戳 (秒)", "美学得分", "清晰度 (Laplacian)", "平均亮度", "去噪帧差 (MSE)",
               "状态/备注"]
    ws.append(headers)

    # 写入数据
    for item in data_list:
        ws.append([
            item["frame_id"],
            item["save_path"],
            item["timestamp_sec"],
            item["aesthetic_score"],
            item["sharpness"],
            item["brightness"],
            item["mse_diff"],
            item["remark"]
        ])

    # 样式配置（优雅冷灰/商务蓝风格）
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    fill_header = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")  # 深冷灰
    fill_zebra = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")  # 浅灰交替隔行

    thin_border = Border(
        left=Side(style='thin', color='E5E7E9'),
        right=Side(style='thin', color='E5E7E9'),
        top=Side(style='thin', color='E5E7E9'),
        bottom=Side(style='thin', color='E5E7E9')
    )

    # 格式化表头
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # 格式化数据行
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)),
                                  start=2):
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)

        for col_idx, cell in enumerate(row, start=1):
            cell.font = font_body
            cell.border = thin_border

            # 斑马纹
            if is_even:
                cell.fill = fill_zebra

            # 数据对齐与数字格式化
            if col_idx in [1, 3]:  # 帧序号、秒数
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx == 3: cell.number_format = '0.00'
            elif col_idx in [4, 5, 6, 7]:  # 指标列
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_idx == 4: cell.number_format = '0.000'
                if col_idx in [5, 6, 7]: cell.number_format = '0.00'
            else:  # 路径和备注
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 对美学得分（第4列）应用条件格式（由浅黄到翠绿的色带）
    color_scale = ColorScaleRule(start_type='num', start_value=4.5, start_color='FFF2CC',
                                 mid_type='num', mid_value=6.0, mid_color='E2EFDA',
                                 end_type='num', end_value=8.0, end_color='C6E0B4')
    ws.conditional_formatting.add(f"D2:D{ws.max_row}", color_scale)

    # 自动调整列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                # 处理中文和长路径的合适显示宽度
                val_str = str(cell.value)
                byte_len = len(val_str.encode('utf-8'))
                max_len = max(max_len, byte_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f" [Excel生成] 报表已成功导出并美化至 -> {output_path}")


# 主抽帧逻辑
def frame_capture_second(interval_sec: float):
    try:
        info = get_video_info(video_path)
        print(f"--- 视频加载成功 ---")
        print(
            f"分辨率: {info['resolution_ratio']} | FPS: {info['fps']} | 总帧数: {info['frame_count']} | 总时长: {info['duration']}秒")
    except Exception as e:
        print(f"视频加载失败: {e}")
        return

    cap = cv2.VideoCapture(video_path)
    fps = info['fps']
    frame_interval = int(fps * interval_sec)

    frame_id = 0
    save_id = 0
    last_saved_frame = None

    # 用于存储需要写入 Excel 的数据
    captured_frames_data = []

    print(f"开始抽样，每 {interval_sec} 秒（即每 {frame_interval} 帧）评估一次...\n")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # 到达抽样步长
        if frame_id % frame_interval == 0:
            current_second = round(frame_id / fps, 2)
            print(f" [评估检测] 正在检查视频第 {frame_id} 帧 (约 {current_second} 秒处):")

            # 进行质量判定并获取各项底层指标
            is_passed, metrics = check_frame_quality(frame, prev_frame=last_saved_frame)

            if is_passed:
                frame_name = f"frame_{save_id:06d}.jpg"
                save_path = os.path.join(output_dir, frame_name)
                cv2.imwrite(save_path, frame)
                print(f"       [成功捕获] 高质量帧已保存至 -> {save_path}")

                # 组装成功帧的数据
                frame_info = {
                    "frame_id": frame_id,
                    "save_path": save_path,
                    "timestamp_sec": current_second,
                    "aesthetic_score": metrics["aesthetic_score"],
                    "sharpness": metrics["sharpness"],
                    "brightness": metrics["brightness"],
                    "mse_diff": metrics["mse_diff"] if metrics["mse_diff"] is not None else 0.0,
                    "remark": metrics["remark"]
                }
                captured_frames_data.append(frame_info)

                save_id += 1
                last_saved_frame = frame.copy()
            print("-" * 50)

        if frame_id % 100 == 0 and frame_id > 0:
            print(f"[系统进度] 已扫描至第 {frame_id}/{info['frame_count']} 帧...")

        frame_id += 1

    cap.release()

    print(f"\n==========================================")
    print(f"  抽帧阶段完成！总共扫描了 {frame_id} 帧，最终截取高质量帧: {save_id} 张。")
    print(f"==========================================")

    # 如果截取到了高质量帧，执行 Excel 导出
    if captured_frames_data:
        print("正在生成结构化 Excel 报表...")
        save_to_styled_excel(captured_frames_data, excel_output_path)
    else:
        print("提示：未拦截到任何满足高质量阈值的视频帧，不生成 Excel。")


# ===================== 新增：批量图片评估接口 =====================

def batch_score_images(
    image_dir: str,
    output_excel: str = "batch_image_quality.xlsx",
    skip_similar: bool = False
):
    """
    批量对文件夹下的图片进行质量评估与美学打分

    :param image_dir: 图片文件夹路径
    :param output_excel: 输出 Excel 文件名
    :param skip_similar: 是否启用相邻图片相似度过滤
    """
    image_dir = Path(image_dir)
    if not image_dir.exists():
        raise FileNotFoundError(f"图片目录不存在: {image_dir}")

    image_paths = sorted(
        [p for p in image_dir.iterdir() if p.suffix.lower() in EXTS]
    )

    if not image_paths:
        print("⚠️ 文件夹下未找到图片")
        return

    print(f"\n{'='*60}")
    print(f"  批量图片质量评估启动")
    print(f"  图片目录 : {image_dir}")
    print(f"  图片数量 : {len(image_paths)}")
    print(f"{'='*60}\n")

    results = []
    prev_frame = None
    prev_frame_path = None

    for idx, img_path in enumerate(image_paths, start=1):
        frame = cv2.imread(str(img_path))
        if frame is None:
            print(f"[跳过] 无法读取: {img_path.name}")
            continue

        print(f"[{idx}/{len(image_paths)}] 正在评估: {img_path.name}")

        # 质量检查
        is_passed, metrics = check_frame_quality(
            frame,
            prev_frame=prev_frame if skip_similar else None
        )

        # 记录结果（无论是否通过）
        record = {
            "frame_id": idx,
            "save_path": str(img_path),
            "timestamp_sec": None,
            "aesthetic_score": metrics["aesthetic_score"],
            "sharpness": metrics["sharpness"],
            "brightness": metrics["brightness"],
            "mse_diff": metrics["mse_diff"] if metrics["mse_diff"] is not None else 0.0,
            "remark": metrics["remark"]
        }
        results.append(record)

        if is_passed:
            prev_frame = frame.copy()
            prev_frame_path = img_path

        print("-" * 50)

    print(f"\n✅ 评估完成，共处理 {len(results)} 张图片")

    if results:
        save_to_styled_excel(results, output_excel)
        print(f"📊 质量报告已导出: {output_excel}")
    else:
        print("⚠️ 无有效数据，未生成 Excel")


if __name__ == "__main__":
    img = cv2.imread(r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\动态范围_中\frame_000000_动态范围_中.jpg")
    print(img is None)

    #frame_capture_second(2.5)
    batch_score_images(
        image_dir=r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\dynamic_mid",
        output_excel="frames_1080P_动态范围_中.xlsx",
        skip_similar=True
    )
    batch_score_images(
        image_dir=r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\dynamic_high",
        output_excel="frames_1080P_动态范围_高.xlsx",
        skip_similar=True
    )
    # batch_score_images(
    #     image_dir=r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\noise_mid",
    #     output_excel="frames_1080P_噪点_中.xlsx",
    #     skip_similar=True
    # )
    # batch_score_images(
    #     image_dir=r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\noise_high",
    #     output_excel="frames_1080P_噪点_高.xlsx",
    #     skip_similar=True
    # )


"""
整体评分大约在5.5分左右
1.0 - 3.0 分：废片 / 严重缺陷 (Poor / Defective)
3.0 - 4.5 分：不及格 / 普通快照 (Below Average / Amateur Snapshot)
4.5 - 6.0 分：及格 / 良好 (Average / Good)
6.0 - 7.5 分：优秀 / 专业级 (Very Good / Professional)
7.5 - 9.0 分：卓越 / 艺术级 (Excellent / Artistic)
9.0 - 10.0 分：大师级 / 完美 (Masterpiece / Perfect)
"""