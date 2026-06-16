import os
from pathlib import Path
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

import torch
import torch.nn as nn
import clip
from PIL import Image

# 如果 aibox.test 导入失败，可以取消下面这行的注释作为后备
# EXTS = ('.jpg', '.jpeg', '.png', '.bmp', '.webp')
from aibox.test import EXTS

# 视频路径与输出目录（视频模式使用）
video_path = r"C:\Users\pablozhao\Desktop\Tubedown Download\【大疆行车记录仪】世纪大道，下班高峰[1920x1080]30fps.mp4"
output_dir = "frames_1080P_20260609"
os.makedirs(output_dir, exist_ok=True)
excel_output_path = "high_quality_frames_1080P_v2.xlsx"


# ===================== 1. 美学评分模型定义 =====================

class AestheticPredictorV2(nn.Module):
    def __init__(self, input_dim=768):
        super().__init__()
        self.input_dim = input_dim
        self.layers = nn.Sequential(
            nn.Linear(self.input_dim, 1024),
            nn.Dropout(0.2),
            nn.Linear(1024, 128),
            nn.Dropout(0.2),
            nn.Linear(128, 64),
            nn.Dropout(0.1),
            nn.Linear(64, 16),
            nn.Linear(16, 1)
        )

    def forward(self, x):
        return self.layers(x)


# 初始化模型与设备
WEIGHT_FILE = "ava+logos-l14-linearMSE.pth"
device = "cuda" if torch.cuda.is_available() else "cpu"
print(f"[*] 当前使用设备: {device}")
print("[*] 正在载入 OpenAI CLIP ViT-L/14 ...")
clip_model, clip_preprocess = clip.load("ViT-L/14", device=device)
model_aesthetic = AestheticPredictorV2(768).to(device)
model_aesthetic.load_state_dict(torch.load(WEIGHT_FILE, map_location=device))
model_aesthetic.eval()


def get_laion_v2_score(cv2_frame):
    """ 计算 LAION V2 美学打分 """
    rgb_frame = cv2.cvtColor(cv2_frame, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(rgb_frame)
    image_tensor = clip_preprocess(pil_img).unsqueeze(0).to(device)

    with torch.no_grad():
        image_features = clip_model.encode_image(image_tensor)
        image_features /= image_features.norm(dim=-1, keepdim=True)
        prediction = model_aesthetic(image_features.float())

    return float(prediction.item())


# ===================== 2. 传统图像质量算法 =====================

def get_video_info(video_path: str) -> dict:
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise ValueError(f"无法打开视频文件: {video_path}")
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS)
    frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    resolution_ratio = f"{width}*{height}"

    info = {
        "width": width,
        "height": height,
        "fps": fps,
        "frame_count": frame_count,
        "resolution_ratio": resolution_ratio,
    }
    if info["fps"] > 0:
        info["duration"] = round(info["frame_count"] / info["fps"], 2)
    cap.release()
    return info


def frame_sharpness(frame):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    return cv2.Laplacian(gray, cv2.CV_64F).var()


def frame_diff_roi(img1, img2):
    h, w, _ = img1.shape
    start_y, end_y = int(h * 0.2), int(h * 0.7)
    g1 = cv2.cvtColor(img1[start_y:end_y, :], cv2.COLOR_BGR2GRAY)
    g2 = cv2.cvtColor(img2[start_y:end_y, :], cv2.COLOR_BGR2GRAY)
    return np.mean((g1.astype(float) - g2.astype(float)) ** 2)


def estimate_dynamic_range(img_bgr, percentile_low=1, percentile_high=99) -> float:
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    p_low = np.percentile(gray, percentile_low)
    p_high = np.percentile(gray, percentile_high)
    return float(p_high - p_low)


def estimate_noise_std(img_bgr):
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY).astype(np.float32)
    lap = cv2.Laplacian(gray, cv2.CV_32F)
    edge_strength = np.abs(lap)

    smooth_mask = edge_strength < np.percentile(edge_strength, 30)
    if np.sum(smooth_mask) < 100:
        return float(np.std(gray))

    noise_map = cv2.GaussianBlur(gray, (3, 3), 0)
    noise_residual = gray - noise_map
    return float(np.std(noise_residual[smooth_mask]))


# ===================== 3. 核心质量综合评估 =====================

def check_frame_quality(frame, prev_frame=None):
    """
    综合评估单张图片的各项质量指标
    返回: (是否通过标志位 bool, 包含所有指标的字典 dict)
    """
    metrics = {
        "sharpness": 0.0,
        "brightness": 0.0,
        "dynamic_range": 0.0,
        "noise_std": 0.0,
        "mse_diff": 0.0,
        "aesthetic_score": 0.0,
        "remark": "通过验证"
    }

    # 1. 检查清晰度
    sharp = frame_sharpness(frame)
    metrics["sharpness"] = round(sharp, 2)
    if sharp < 120:
        metrics["remark"] = f"未通过: 清晰度过低 ({sharp:.1f} < 120)"
        return False, metrics

    # 2. 检查亮度
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    brightness = float(gray.mean())
    metrics["brightness"] = round(brightness, 2)
    if brightness < 40 or brightness > 220:
        metrics["remark"] = f"未通过: 曝光异常 (亮度: {brightness:.1f})"
        return False, metrics

    # 3. 检查动态范围
    # dr = estimate_dynamic_range(frame)
    # metrics["dynamic_range"] = round(dr, 2)
    # if dr < 175 or dr > 245:
    #     metrics["remark"] = f"未通过: 动态范围不佳 (HDR: {dr:.1f})"
    #     return False, metrics
    #
    # # 4. 检查噪点强度
    # noise = estimate_noise_std(frame)
    # metrics["noise_std"] = round(noise, 2)
    # if noise > 5.0:
    #     metrics["remark"] = f"未通过: 噪点过高 (Noise: {noise:.2f})"
    #     return False, metrics

    # 5. 画面重复度过滤（仅在提供前一帧时触发）
    if prev_frame is not None:
        diff = frame_diff_roi(prev_frame, frame)
        metrics["mse_diff"] = round(diff, 2)
        if diff < 1500:
            metrics["remark"] = f"未通过: 与前图高度相似 (MSE: {diff:.1f})"
            return False, metrics

    # 6. 美学评分
    aesthetic_score = get_laion_v2_score(frame)
    metrics["aesthetic_score"] = round(aesthetic_score, 3)

    if aesthetic_score >= 7.5:
        metrics["remark"] = "卓越/艺术级"
    elif aesthetic_score >= 6.0:
        metrics["remark"] = "优秀/专业级"
    else:
        metrics["remark"] = "常规高质量"

    return True, metrics


# ===================== 4. Excel 报表生成与美化 =====================

def save_to_styled_excel(data_list, output_path, mode="image"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "质量评估报告"
    ws.views.sheetView[0].showGridLines = True

    # 动态表头（自适应视频模式和纯图片模式）
    if mode == "video":
        headers = ["帧序号", "保存路径", "时间戳 (秒)", "美学得分", "清晰度", "平均亮度", "动态范围", "噪点标准差",
                   "相似帧差(MSE)", "最终状态/备注"]
    else:
        headers = ["图片序号", "图片路径", "美学得分", "清晰度", "平均亮度", "动态范围", "噪点标准差", "相似度(MSE)",
                   "最终状态/备注"]

    ws.append(headers)

    # 写入数据
    for item in data_list:
        if mode == "video":
            row_data = [item["frame_id"], item["save_path"], item["timestamp_sec"], item["aesthetic_score"],
                        item["sharpness"], item["brightness"], item["dynamic_range"], item["noise_std"],
                        item["mse_diff"], item["remark"]]
        else:
            row_data = [item["frame_id"], item["save_path"], item["aesthetic_score"], item["sharpness"],
                        item["brightness"], item["dynamic_range"], item["noise_std"], item["mse_diff"], item["remark"]]
        ws.append(row_data)

    # 样式设计
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")  # 商务深蓝灰
    fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='E5E7E9'), right=Side(style='thin', color='E5E7E9'),
        top=Side(style='thin', color='E5E7E9'), bottom=Side(style='thin', color='E5E7E9')
    )

    # 格式化表头
    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # 判断美学得分所在的列索引
    score_col_idx = 4 if mode == "video" else 3
    score_col_letter = openpyxl.utils.get_column_letter(score_col_idx)

    # 格式化数据行
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)),
                                  start=2):
        ws.row_dimensions[row_idx].height = 22
        is_even = (row_idx % 2 == 0)

        for col_idx, cell in enumerate(row, start=1):
            cell.font = font_body
            cell.border = thin_border
            if is_even:
                cell.fill = fill_zebra

            # 对齐与数字格式微调
            if col_idx == 1:  # 序号
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == score_col_idx:  # 美学得分
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.000'
            elif col_idx in range(score_col_idx + 1, len(headers)):  # 各类数值指标
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.00'
            else:  # 路径与备注说明
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 对美学得分添加条件格式（软色调：浅黄 -> 翠绿）
    color_scale = ColorScaleRule(start_type='num', start_value=4.0, start_color='FFF2CC',
                                 mid_type='num', mid_value=6.0, mid_color='E2EFDA',
                                 end_type='num', end_value=8.0, end_color='C6E0B4')
    ws.conditional_formatting.add(f"{score_col_letter}2:{score_col_letter}{ws.max_row}", color_scale)

    # 自适应计算理想列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                byte_len = len(str(cell.value).encode('utf-8'))
                max_len = max(max_len, byte_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"📊 [报表成功导出] -> {output_path}")


# ===================== 5. 核心业务主逻辑：视频动态抽帧 =====================

def frame_capture_second(interval_sec: float):
    """
    按设定的秒数步长扫描视频，过滤出高质量核心帧，并将其信息保存到 Excel。
    """
    try:
        info = get_video_info(video_path)
        print(f"\n--- 🎬 视频加载成功 ---")
        print(
            f"分辨率: {info['resolution_ratio']} | FPS: {info['fps']} | 总帧数: {info['frame_count']} | 总时长: {info['duration']}秒")
    except Exception as e:
        print(f"❌ 视频加载失败: {e}")
        return

    cap = cv2.VideoCapture(video_path)
    fps = info['fps']
    frame_interval = int(fps * interval_sec)

    frame_id = 0
    save_id = 0
    last_saved_frame = None
    captured_frames_data = []

    print(f"开始抽样，每 {interval_sec} 秒（即每 {frame_interval} 帧）评估一次...\n")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # 到达抽样步长
        if frame_id % frame_interval == 0:
            current_second = round(frame_id / fps, 2)
            print(f" 🔍 [评估检测] 正在检查视频第 {frame_id} 帧 (约 {current_second} 秒处):")

            # 进行一站式质量与美学判定
            is_passed, metrics = check_frame_quality(frame, prev_frame=last_saved_frame)

            if is_passed:
                frame_name = f"frame_{save_id:06d}.jpg"
                save_path = os.path.join(output_dir, frame_name)
                cv2.imwrite(save_path, frame)
                print(f"       ✨ [成功捕获] 高质量帧已保存 -> {save_path} | 美学分: {metrics['aesthetic_score']}")

                # 组装成功帧的数据（注意：这里完整收集了动态范围与噪点）
                frame_info = {
                    "frame_id": frame_id,
                    "save_path": save_path,
                    "timestamp_sec": current_second,
                    "aesthetic_score": metrics["aesthetic_score"],
                    "sharpness": metrics["sharpness"],
                    "brightness": metrics["brightness"],
                    "dynamic_range": metrics["dynamic_range"],
                    "noise_std": metrics["noise_std"],
                    "mse_diff": metrics["mse_diff"],
                    "remark": metrics["remark"]
                }
                captured_frames_data.append(frame_info)
                save_id += 1
                last_saved_frame = frame.copy()
            else:
                print(f"       ❌ [拦截未通过] 原因: {metrics['remark']}")
            print("-" * 50)

        if frame_id % 500 == 0 and frame_id > 0:
            print(f" ⏳ [系统进度] 已扫描至第 {frame_id}/{info['frame_count']} 帧...")

        frame_id += 1

    cap.release()

    print(f"\n==========================================")
    print(f" 🎉 抽帧阶段完成！总扫描 {frame_id} 帧，最终拦截高质量帧: {save_id} 张。")
    print(f"==========================================")

    if captured_frames_data:
        print("正在生成结构化 Excel 报表...")
        save_to_styled_excel(captured_frames_data, excel_output_path, mode="video")
    else:
        print("提示：未拦截到任何满足高质量阈值的视频帧，放弃生成 Excel。")


# ===================== 6. 核心业务主逻辑：批量图片评估 =====================

def batch_score_images(image_dir: str, output_excel: str = "batch_image_quality.xlsx", skip_similar: bool = False):
    """
    批量评估指定文件夹下的所有图片，并将所有质量明细（包含通过与未通过）生成 Excel 报告。
    """
    image_dir = Path(image_dir)
    if not image_dir.exists():
        raise FileNotFoundError(f"指定的图片目录不存在: {image_dir}")

    image_paths = sorted([p for p in image_dir.iterdir() if p.suffix.lower() in EXTS])

    if not image_paths:
        print(f"⚠️  在目录 {image_dir} 下未找到匹配的图片格式 {EXTS}")
        return

    print("\n" + "=" * 60)
    print(f"  🚀 批量图片质量评估引擎启动")
    print(f"  📂 目标目录 : {image_dir}")
    print(f"  📸 图片总量 : {len(image_paths)}")
    print("=" * 60 + "\n")

    results = []
    prev_frame = None

    for idx, img_path in enumerate(image_paths, start=1):
        frame = cv2.imread(str(img_path))
        if frame is None:
            print(f"[-] [{idx}/{len(image_paths)}] 无法读取(跳过): {img_path.name}")
            continue

        is_passed, metrics = check_frame_quality(frame, prev_frame=prev_frame if skip_similar else None)

        record = {
            "frame_id": idx,
            "save_path": str(img_path),
            "aesthetic_score": metrics["aesthetic_score"],
            "sharpness": metrics["sharpness"],
            "brightness": metrics["brightness"],
            "dynamic_range": metrics["dynamic_range"],
            "noise_std": metrics["noise_std"],
            "mse_diff": metrics["mse_diff"],
            "remark": metrics["remark"]
        }
        results.append(record)

        status_str = "[Pass]" if is_passed else "[Fail]"
        print(
            f"[{idx}/{len(image_paths)}] {status_str} {img_path.name} | Score: {metrics['aesthetic_score']:.2f} | {metrics['remark']}")

        if is_passed and skip_similar:
            prev_frame = frame.copy()

    if results:
        save_to_styled_excel(results, output_excel, mode="image")
    else:
        print("⚠️ 未产生任何评估数据。")


# ===================== 7. 运行入口 =====================

if __name__ == "__main__":
    # ---- 任务 1：视频抽帧（根据需要取消注释运行） ----
    #frame_capture_second(5)

    # ---- 任务 2：批量图片评估 ----
    batch_score_images(
        image_dir=r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\dynamic_mid",
        output_excel="frames_1080P_动态范围_中.xlsx",
        skip_similar=True
    )

    batch_score_images(
        image_dir=r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\dynamic_high",
        output_excel="frames_1080P_动态范围_高.xlsx",
        skip_similar=True
    )
    batch_score_images(
        image_dir=r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\noise_mid",
        output_excel="frames_1080P_噪点_中.xlsx",
        skip_similar=True
    )
    batch_score_images(
        image_dir=r"C:\pablo\05_self_code\yolo-sam-lama\src\aibox\degrade_output\noise_high",
        output_excel="frames_1080P_噪点_高.xlsx",
        skip_similar=True
    )